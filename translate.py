import sys, os
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from datetime import datetime

try: 
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.utils.cell import coordinate_from_string
except ImportError:
    from openpyxl.utils import get_column_letter, coordinate_from_string, column_index_from_string


# Python 3 版本校验(必须使用 Python 3)
if sys.version_info[0] < 3:
    raise "Must be using Python 3"

# print ('Number of arguments:', len(sys.argv), 'arguments.')
# print ('Argument List:', str(sys.argv))

if len(sys.argv) < 2 :
    print('输入异常! 需要您输入翻译的文件(.xlsx格式)!')
    sys.exit(0)

# 加载xlsx文件
filename = sys.argv[1]
wb = load_workbook(filename = filename)
sheet = wb.active 

# 翻译的索引, 默认: A列
translateColumnIndex = 'A'
# 中文的索引, 默认: B列
zhHansIndex = 'B'

if len(sys.argv) > 2 :
    translateColumnIndex = sys.argv[2]
print('当前工作表的英文内容列:' + translateColumnIndex)

if len(sys.argv) > 3 :
    zhHansIndex = sys.argv[3]
print('当前工作表的中文内容列:' + zhHansIndex)


#忽略 行 
ignore_row_number = 0
if len(sys.argv) > 4:
    ignore_row_number = int(sys.argv[4])
print('忽略 行:' + str(ignore_row_number))

# 忽略 列 
ignore_column_number = 0
if len(sys.argv) > 5:
    ignore_column_number = int(sys.argv[5])
print('忽略 列:' + str(ignore_column_number))

# 获取文件夹路径
folder = os.path.dirname(os.path.abspath(filename))
print('工作文件夹路径: '+ folder)

# 读取行及列
row_count = sheet.max_row
column_count = sheet.max_column
invalid_cell_count = 0
print("共："+str(row_count)+'行'+'   '+str(column_count)+'列')
# print ('File: ' + filename  + " -> " + str(row_count) + " rows and " + str(sheet.get_highest_column()) + " columns")

# 遍历
for column in sheet.columns: 
    # 目标语言 列索引 (默认：每列第一个元素。如：A1, A2, A3)
    targetLanguageColumn = column[0].column
        
    # 第0行第X列是否为空
    if column[0].value:
    # 当列第X行是否为空
    # if sheet[str(targetLanguageColumn+1)][0].value:    #check if there is a language name (avoid empty)
        # X 列第一个元素，
        if column[0].value:
            outputFilename = column[0].value + '.strings'
        else:
            outputFilename = "NULL" + '.strings'
        print('文件名: ' + outputFilename)
        outputFolder = folder + '/'+filename+'-output/'
        
        if not os.path.exists(outputFolder):
            os.makedirs(outputFolder)

        dotStringFile = open(outputFolder + outputFilename, 'w', encoding='utf8')
        print('生成文件: ' + outputFolder + outputFilename)

        # 生成 .strings 文件内容
        fileContent = '/* Auto generated .string file from EXCEL with Python script. \nAuthor: raxcat@github \nCopyright: raxcat@github 2015 \nTimestamp:'+ str(datetime.now())+'*/\n\n'

        for cell in column:
            if cell.value :  #check if there is any value(not empty cell)
                # column_index_from_string()
                # print (cell.column)

                if cell.row > ignore_row_number or cell.column >  ignore_column_number: #check ignore row and column

                    languageIndex = str(cell.column)+"1"
                    language = (str(sheet[languageIndex][0].value)).strip()

                    # print ( '[' + str(cell.column)+str(cell.row)+']'  + '(' +language + ')')

                    annotationContentIndex = translateColumnIndex + str(cell.row)
                    annotationContent = (str(sheet[annotationContentIndex].value)).rstrip()   #string due to some err in excel
                    translatedContent = str(cell.value).rstrip()

                    zhHansContentIndex = zhHansIndex + str(cell.row)
                    zhHansContent = (str(sheet[zhHansContentIndex].value)).rstrip()   #string due to some err in excel
                    # translatedContent = str(cell.value).rstrip()

                    # 注释内容，如：/* */
                    line1 = '/* ' + annotationContent + ' : ' + zhHansContent  + ' */'
                    # 翻译的主要内容
                    line2 = '"'+annotationContent + '" = "' + translatedContent + '";'

                    fileContent += line1+'\n'
                    fileContent += line2+'\n'
                    fileContent += ('\n')
            else:
                invalid_cell_count += 1

        dotStringFile.write(fileContent)

    else:
        invalid_cell_count += 1

print('There are ' + str(invalid_cell_count) + ' empty cells in raw excel file')

# Thanks: @raxcat
# https://github.com/raxcat/ExcelToStrings_ObjC